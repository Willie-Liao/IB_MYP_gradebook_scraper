import openpyxl
wb = openpyxl.load_workbook('gradebook_class_11423945_unknown_term_20260120_225037.xlsx')
sheet = wb.active
print('Headers:')
for i, cell in enumerate(sheet[1], 1):
    print(f'  {i}. {cell.value}')
print(f'\nTotal columns: {sheet.max_column}')
print(f'Total rows: {sheet.max_row}')
print('\nFirst student row (sample):')
for i, cell in enumerate(sheet[2], 1):
    val = str(cell.value) if cell.value else ''
    if val:
        print(f'  Col {i}: {val[:80]}')
