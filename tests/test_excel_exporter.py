"""Property tests for ExcelExporter.

Feature: managebac-gradebook-scraper, Property 5: Excel Export Round-Trip
Validates: Requirements 6.1, 6.2, 6.3, 6.4, 6.5, 6.6
"""

import os
import tempfile
from typing import Any

from hypothesis import given, settings, strategies as st
from openpyxl import load_workbook

from src.excel_exporter import ExcelExporter
from src.models import GradebookData, Score, Student, Task, TermGrade


# Strategy to generate valid student data
@st.composite
def student_strategy(draw: st.DrawFn) -> Student:
    """Generate a random Student object."""
    student_id = draw(st.text(
        alphabet=st.characters(whitelist_categories=("Nd", "Lu", "Ll")),
        min_size=1,
        max_size=10
    ))
    student_name = draw(st.text(
        alphabet=st.characters(whitelist_categories=("Lu", "Ll", "Zs")),
        min_size=1,
        max_size=30
    ).filter(lambda x: x.strip()))
    return Student(id=student_id, name=student_name.strip())


@st.composite
def task_strategy(draw: st.DrawFn) -> Task:
    """Generate a random Task object."""
    task_id = draw(st.text(
        alphabet=st.characters(whitelist_categories=("Nd", "Lu", "Ll")),
        min_size=1,
        max_size=10
    ))
    task_name = draw(st.text(
        alphabet=st.characters(whitelist_categories=("Lu", "Ll", "Zs")),
        min_size=1,
        max_size=30
    ).filter(lambda x: x.strip()))
    task_link = f"/tasks/{task_id}"
    return Task(id=task_id, name=task_name.strip(), link=task_link)


@st.composite
def score_strategy(
    draw: st.DrawFn, 
    student_ids: list[str], 
    task_ids: list[str]
) -> Score:
    """Generate a random Score object for given students and tasks."""
    student_id = draw(st.sampled_from(student_ids))
    task_id = draw(st.sampled_from(task_ids))
    criterion = draw(st.sampled_from(["A", "B", "C", "D"]))
    score = draw(st.one_of(
        st.none(),
        st.integers(min_value=0, max_value=8)
    ))
    comment = draw(st.one_of(
        st.none(),
        st.text(
            alphabet=st.characters(whitelist_categories=("Lu", "Ll", "Zs", "Nd")),
            min_size=1,
            max_size=50
        ).filter(lambda x: x.strip())
    ))
    return Score(
        student_id=student_id,
        task_id=task_id,
        criterion=criterion,
        score=score,
        comment=comment.strip() if comment else None
    )


@st.composite
def term_grade_strategy(draw: st.DrawFn, student_names: list[str]) -> TermGrade:
    """Generate a random TermGrade object for given student names."""
    student_name = draw(st.sampled_from(student_names))
    grade = draw(st.sampled_from(["1", "2", "3", "4", "5", "6", "7", "8", "INC", "N/A"]))
    return TermGrade(student_name=student_name, grade=grade)


@st.composite
def gradebook_data_strategy(draw: st.DrawFn) -> GradebookData:
    """Generate a random GradebookData object."""
    # Generate students (at least 1)
    students = draw(st.lists(
        student_strategy(),
        min_size=1,
        max_size=10,
        unique_by=lambda x: x.name,
    ))
    
    # Generate tasks (at least 1)
    tasks = draw(st.lists(
        task_strategy(),
        min_size=1,
        max_size=5,
        unique_by=lambda x: x.id
    ))
    
    # Generate scores for student-task combinations
    student_ids = [s.id for s in students]
    task_ids = [t.id for t in tasks]
    
    scores: list[Score] = []
    # Generate some scores (not necessarily for all combinations)
    num_scores = draw(st.integers(min_value=1, max_value=len(students) * len(tasks) * 2))
    for _ in range(num_scores):
        score = draw(score_strategy(student_ids, task_ids))
        # Avoid duplicate student-task-criterion combinations
        key = (score.student_id, score.task_id, score.criterion)
        if not any((s.student_id, s.task_id, s.criterion) == key for s in scores):
            scores.append(score)
    
    # Generate term grades for students
    student_names = [s.name for s in students]
    term_grades: list[TermGrade] = []
    for name in student_names:
        grade = draw(st.sampled_from(["1", "2", "3", "4", "5", "6", "7", "8", "INC", "N/A"]))
        term_grades.append(TermGrade(student_name=name, grade=grade))
    
    # Generate class and term names
    class_name = draw(st.text(
        alphabet=st.characters(whitelist_categories=("Lu", "Ll", "Nd", "Zs")),
        min_size=1,
        max_size=20
    ).filter(lambda x: x.strip()))
    
    term_name = draw(st.text(
        alphabet=st.characters(whitelist_categories=("Lu", "Ll", "Nd", "Zs")),
        min_size=1,
        max_size=20
    ).filter(lambda x: x.strip()))
    
    return GradebookData(
        students=students,
        tasks=tasks,
        scores=scores,
        term_grades=term_grades,
        class_name=class_name.strip(),
        term_name=term_name.strip()
    )


def read_excel_data(filepath: str) -> dict[str, Any]:
    """Read data from an Excel file for verification.
    
    Returns a dictionary with:
    - headers: list of column headers
    - rows: list of dictionaries mapping header to cell value
    - comments: dict mapping (row, col) to comment text
    """
    workbook = load_workbook(filepath)
    sheet = workbook.active
    
    if sheet is None:
        return {"headers": [], "rows": [], "comments": {}}
    
    # Read headers from first row
    headers: list[str] = []
    for cell in sheet[1]:
        headers.append(str(cell.value) if cell.value else "")
    
    # Read data rows
    rows: list[dict[str, Any]] = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        row_data: dict[str, Any] = {}
        for col_idx, cell in enumerate(row):
            header = headers[col_idx] if col_idx < len(headers) else f"col_{col_idx}"
            row_data[header] = cell.value
        rows.append(row_data)
    
    # Read comments
    comments: dict[tuple[int, int], str] = {}
    for row_idx, row in enumerate(sheet.iter_rows(), start=1):
        for col_idx, cell in enumerate(row, start=1):
            if cell.comment:
                comments[(row_idx, col_idx)] = cell.comment.text
    
    return {"headers": headers, "rows": rows, "comments": comments}


# Feature: managebac-gradebook-scraper, Property 5: Excel Export Round-Trip
# Validates: Requirements 6.1, 6.2, 6.3, 6.4, 6.5, 6.6
@given(data=gradebook_data_strategy())
@settings(max_examples=100)
def test_excel_export_round_trip(data: GradebookData) -> None:
    """Property test: Excel export preserves all gradebook data.
    
    For any valid GradebookData object, exporting to Excel and reading back
    SHALL produce:
    - Student names in the first column matching the original data
    - Task columns with correct headers
    - Score cells in "[Criterion]: [Score]" format
    - Term grades in the last column
    - Comments preserved in cell notes
    """
    # Create a temporary file for the export
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        output_path = tmp.name
    
    try:
        # Export the data
        ExcelExporter.export(data, output_path)
        
        # Read back the exported data
        excel_data = read_excel_data(output_path)
        headers = excel_data["headers"]
        rows = excel_data["rows"]
        
        # Property 1: First column is "Student Name" (Requirement 6.1)
        assert headers[0] == "Student Name", (
            f"First column should be 'Student Name', got '{headers[0]}'"
        )
        
        # Property 2: Term Grade column exists (Requirement 6.3)
        # Note: Additional evaluation columns come after Term Grade
        assert "Term Grade" in headers, (
            f"'Term Grade' column should exist in headers"
        )
        term_grade_idx = headers.index("Term Grade")
        
        # Property 3: Number of data rows equals number of students
        assert len(rows) == len(data.students), (
            f"Expected {len(data.students)} rows, got {len(rows)}"
        )
        
        # Property 4: Student names are preserved (Requirement 6.1)
        exported_names = [row.get("Student Name") for row in rows]
        expected_names = [s.name for s in data.students]
        assert exported_names == expected_names, (
            f"Student names mismatch: expected {expected_names}, got {exported_names}"
        )
        
        # Property 5: Term grades are preserved (Requirement 6.3)
        term_grades_lookup = {tg.student_name: tg.grade for tg in data.term_grades}
        for row_idx, row in enumerate(rows):
            student_name = row.get("Student Name")
            expected_grade = term_grades_lookup.get(student_name, "")
            actual_grade = row.get("Term Grade", "")
            assert actual_grade == expected_grade, (
                f"Term grade mismatch for '{student_name}': "
                f"expected '{expected_grade}', got '{actual_grade}'"
            )
        
        # Property 6: Scores are formatted correctly (Requirement 6.4)
        scores_lookup = {
            (s.student_id, s.task_id, s.criterion): s 
            for s in data.scores
        }
        
        for row_idx, row in enumerate(rows):
            student = data.students[row_idx]
            for task in data.tasks:
                # Check each criterion column for this task
                for criterion in ["A", "B", "C", "D"]:
                    score = scores_lookup.get((student.id, task.id, criterion))
                    if score:
                        header = f"{task.name} ({criterion})"
                        if header in row:
                            cell_value = row[header]
                            if cell_value:
                                expected_value = score.score if score.score is not None else "N/A"
                                expected_format = f"{criterion}: {expected_value}"
                                assert cell_value == expected_format, (
                                    f"Score format mismatch: expected '{expected_format}', "
                                    f"got '{cell_value}'"
                                )
        
    finally:
        # Clean up the temporary file
        if os.path.exists(output_path):
            os.remove(output_path)


# Feature: managebac-gradebook-scraper, Property 5: Excel Export Round-Trip
# Validates: Requirements 6.5
@given(data=gradebook_data_strategy())
@settings(max_examples=100)
def test_excel_export_preserves_comments(data: GradebookData) -> None:
    """Property test: Comments are preserved in adjacent comment columns."""
    # Create a temporary file for the export
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        output_path = tmp.name
    
    try:
        # Export the data
        ExcelExporter.export(data, output_path)
        
        # Read back the exported data including comments
        excel_data = read_excel_data(output_path)
        headers = excel_data["headers"]
        rows = excel_data["rows"]
        scores_with_comments = [s for s in data.scores if s.comment]

        for score in scores_with_comments:
            student = next((s for s in data.students if s.id == score.student_id), None)
            task = next((t for t in data.tasks if t.id == score.task_id), None)
            if student is None or task is None:
                continue

            comment_header = f"{task.name} ({score.criterion}) Comment"
            if comment_header not in headers:
                continue

            row = next((r for r in rows if r.get("Student Name") == student.name), None)
            if row is None:
                continue

            assert row.get(comment_header) == score.comment, (
                f"Comment mismatch for {student.name} / {comment_header}"
            )
        
    finally:
        # Clean up the temporary file
        if os.path.exists(output_path):
            os.remove(output_path)


# Feature: managebac-gradebook-scraper, Property 5: Excel Export Round-Trip
# Validates: Requirements 6.6
def test_excel_export_generates_descriptive_filename() -> None:
    """Test that export generates descriptive filename with class and term info."""
    data = GradebookData(
        students=[Student(id="1", name="Test Student")],
        tasks=[Task(id="1", name="Test Task", link="/tasks/1")],
        scores=[Score(student_id="1", task_id="1", criterion="A", score=7, comment=None)],
        term_grades=[TermGrade(student_name="Test Student", grade="7")],
        class_name="Math 10A",
        term_name="Term 1"
    )
    
    # Export without specifying output path
    output_path = ExcelExporter.export(data)
    
    try:
        # Verify filename contains class and term info
        assert "Math_10A" in output_path or "Math10A" in output_path, (
            f"Filename should contain class name: {output_path}"
        )
        assert "Term_1" in output_path or "Term1" in output_path, (
            f"Filename should contain term name: {output_path}"
        )
        assert output_path.endswith(".xlsx"), (
            f"Filename should end with .xlsx: {output_path}"
        )
        
        # Verify file was created
        assert os.path.exists(output_path), f"File was not created: {output_path}"
        
    finally:
        # Clean up
        if os.path.exists(output_path):
            os.remove(output_path)
