"""Excel export module for ManageBac Gradebook Scraper."""

import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import GradebookData, Score, Task, TermGrade

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Exports gradebook data to Excel files."""

    STUDENT_COLUMN_HEADER: str = "Student Name"
    TERM_GRADE_COLUMN_HEADER: str = "Term Grade"
    CRITERION_COLUMNS: list[str] = [
        "Criterion A",
        "Criterion B",
        "Criterion C",
        "Criterion D",
    ]

    @staticmethod
    def export(data: GradebookData, output_path: str | None = None) -> str:
        workbook = Workbook()
        sheet: Any = workbook.active
        if sheet is None:
            sheet = workbook.create_sheet("Gradebook")
        else:
            sheet.title = "Gradebook"

        columns = ExcelExporter._build_column_structure(data.tasks, data.scores)
        ExcelExporter._write_header_row(sheet, columns)
        ExcelExporter._write_student_rows(sheet, data, columns)
        ExcelExporter._adjust_column_widths(sheet)

        if not output_path:
            output_path = ExcelExporter._generate_filename(data.class_name, data.term_name)

        workbook.save(output_path)
        logger.info(f"Exported gradebook data to: {output_path}")
        return output_path

    @staticmethod
    def _build_column_structure(
        tasks: list[Task],
        scores: list[Score],
    ) -> list[tuple[str, str | None, str | None, bool]]:
        columns: list[tuple[str, str | None, str | None, bool]] = []
        columns.append((ExcelExporter.STUDENT_COLUMN_HEADER, None, None, False))
        columns.append((ExcelExporter.TERM_GRADE_COLUMN_HEADER, None, None, False))

        for header in ExcelExporter.CRITERION_COLUMNS:
            columns.append((header, None, None, False))

        task_criteria: dict[str, set[str]] = {}
        for score in scores:
            task_criteria.setdefault(score.task_id, set()).add(score.criterion)

        for task in tasks:
            criteria = sorted(task_criteria.get(task.id, set()))
            if not criteria:
                columns.append((task.name, task.id, None, False))
                columns.append((f"{task.name} Comment", task.id, None, True))
            else:
                for criterion in criteria:
                    columns.append((f"{task.name} ({criterion})", task.id, criterion, False))
                    columns.append(
                        (f"{task.name} ({criterion}) Comment", task.id, criterion, True)
                    )

        return columns

    @staticmethod
    def _write_header_row(
        sheet: Worksheet,
        columns: list[tuple[str, str | None, str | None, bool]],
    ) -> None:
        for col_idx, (header, _, _, _) in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    @staticmethod
    def _write_student_rows(
        sheet: Worksheet,
        data: GradebookData,
        columns: list[tuple[str, str | None, str | None, bool]],
    ) -> None:
        scores_lookup = ExcelExporter._build_scores_lookup(data.scores)
        term_lookup: dict[str, TermGrade] = {
            tg.student_name: tg for tg in data.term_grades
        }

        for row_idx, student in enumerate(data.students, start=2):
            term_profile = term_lookup.get(student.name)
            for col_idx, (header, task_id, criterion, is_comment) in enumerate(
                columns, start=1
            ):
                if header == ExcelExporter.STUDENT_COLUMN_HEADER:
                    sheet.cell(row=row_idx, column=col_idx, value=student.name)
                elif header == ExcelExporter.TERM_GRADE_COLUMN_HEADER:
                    sheet.cell(
                        row=row_idx,
                        column=col_idx,
                        value=term_profile.grade if term_profile else "",
                    )
                elif header == "Criterion A":
                    sheet.cell(
                        row=row_idx,
                        column=col_idx,
                        value=term_profile.criterion_a if term_profile else "",
                    )
                elif header == "Criterion B":
                    sheet.cell(
                        row=row_idx,
                        column=col_idx,
                        value=term_profile.criterion_b if term_profile else "",
                    )
                elif header == "Criterion C":
                    sheet.cell(
                        row=row_idx,
                        column=col_idx,
                        value=term_profile.criterion_c if term_profile else "",
                    )
                elif header == "Criterion D":
                    sheet.cell(
                        row=row_idx,
                        column=col_idx,
                        value=term_profile.criterion_d if term_profile else "",
                    )
                elif task_id and criterion and is_comment:
                    score = scores_lookup.get((student.id, task_id, criterion))
                    if score and score.comment:
                        sheet.cell(row=row_idx, column=col_idx, value=score.comment)
                elif task_id and criterion and not is_comment:
                    score = scores_lookup.get((student.id, task_id, criterion))
                    if score:
                        sheet.cell(
                            row=row_idx,
                            column=col_idx,
                            value=ExcelExporter._format_score(score),
                        )
                elif task_id and is_comment:
                    for key, score in scores_lookup.items():
                        if key[0] == student.id and key[1] == task_id and score.comment:
                            sheet.cell(row=row_idx, column=col_idx, value=score.comment)
                            break
                elif task_id and not is_comment:
                    for key, score in scores_lookup.items():
                        if key[0] == student.id and key[1] == task_id:
                            sheet.cell(
                                row=row_idx,
                                column=col_idx,
                                value=ExcelExporter._format_score(score),
                            )
                            break

    @staticmethod
    def _build_scores_lookup(scores: list[Score]) -> dict[tuple[str, str, str], Score]:
        lookup: dict[tuple[str, str, str], Score] = {}
        for score in scores:
            lookup[(score.student_id, score.task_id, score.criterion)] = score
        return lookup

    @staticmethod
    def _format_score(score: Score) -> str:
        score_value = str(score.score) if score.score is not None else "N/A"
        return f"{score.criterion}: {score_value}"

    @staticmethod
    def _adjust_column_widths(sheet: Worksheet) -> None:
        for column_cells in sheet.columns:
            max_length = 0
            first_cell = column_cells[0]
            if first_cell.column is None:
                continue
            column_letter = get_column_letter(first_cell.column)
            for cell in column_cells:
                cell_value = str(cell.value) if cell.value else ""
                max_length = max(max_length, len(cell_value))
            sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 50)

    @staticmethod
    def _generate_filename(class_name: str, term_name: str) -> str:
        gradebook_dir = Path("gradebook")
        gradebook_dir.mkdir(exist_ok=True)
        safe_class = ExcelExporter._sanitize_filename(class_name)
        safe_term = ExcelExporter._sanitize_filename(term_name)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return str(gradebook_dir / f"gradebook_{safe_class}_{safe_term}_{timestamp}.xlsx")

    @staticmethod
    def _sanitize_filename(name: str) -> str:
        if not name:
            return "unknown"
        sanitized = re.sub(r"[^\w\-]", "_", name)
        sanitized = re.sub(r"_+", "_", sanitized).strip("_")
        return sanitized or "unknown"
